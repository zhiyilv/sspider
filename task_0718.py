# import modules
# requests, beautifulsoup, re, time
import requests
import time
from bs4 import BeautifulSoup as bs
import re
from pprint import pprint
import openpyxl
import os
# from itertools import chain
# import pandas
from selenium import webdriver


start_url = 'http://ssh.speiyou.com/search/index/grade:1/subject:/level:bx/' \
            'term:/period:/teaid:/m:/d:/time:/bg:n/nu:/service:/curpage:1'
# get all urls for different grades, store the urls into a list
# grade_urls = []
start_soup = bs(requests.get(start_url).content, 'lxml')
search_panel = start_soup.find(id='search-term')
grade_a = search_panel.find('dl')('a')[4:]
domain = 'http://ssh.speiyou.com'
grade_urls = [domain + i.get('href') for i in grade_a]

#-----excel settings--------
excel_name = 'speiyou_shanghai.xlsx'
if excel_name not in os.listdir('.//'):
    my_excel = openpyxl.Workbook()
else:
    my_excel = openpyxl.load_workbook(excel_name)
strip_prefix = '上海'  # sheet title exclude this
header_row = ['teacher', 'title', 'price', 'satisfaction', 'subject', 'grade', 'duration', 'time', 'location']

# -------------------define methods
# parse_c for parsing a container in page
def parse_c(container):
    # the result is in format [teacher, title, price, satisfaction, subject, grade, duration, time, location]
    teacher = container.find(class_='s-r-list-photo').p.text
    if teacher == '无':
        return None
    title = container.find(class_='s-r-list-info').h3.text
    price = float(container.find(class_='price').span.text)
    try:
        satisfaction = float(container.find(class_='pf').find_all('div')[-2].text[:-1])
    except:
        satisfaction = -1
    info = container(text=re.compile('：'))[1:]
    # subject, grade, opendate, schedule, location = [i.split('：')[-1].strip() for i in info]
    info = [i.split('：')[-1].strip() for i in info]
    return [teacher, title, price, satisfaction, *info]


# parse_page for parsing a page of classes
def parse_page(page_soup):
    containers = page_soup(class_='s-r-list')
    classes = [parse_c(i) for i in containers]
    return [i for i in classes if i]


# pass a grade_url, i.e. the first page of classes of a grade
# loop all pages and parse classes inside
def crawl_grade(grade_url):
    soup = bs(myget(grade_url), 'lxml')

    # get the grade
    grade_name = soup.title.text.split('-')[0].strip().lstrip(strip_prefix)
    print('crawling grade {}'.format(grade_name))

    # get the number of maxpage
    last_url = soup.find('a', text='尾页').get('href')
    maxpage = last_url.split(':')[-1]

    # get the common url from grade_url
    common_url = domain + last_url.rstrip(maxpage)
    # create url_list
    url_list = [common_url + str(i) for i in range(2, int(maxpage)+1)]

    print('there are {} pages to crawl for this grade'.format(maxpage))

    # page_class_list = [parse_page(soup)]
    first_page = parse_page(soup)
    print('found {} valid classes on first page'.format(len(first_page)))
    append_to_sheet(first_page, grade_name)
    # show(first_page)

    count = len(first_page)
    for url in url_list[:0]:
        print('in page: {}'.format(url))
        page_soup = bs(myget(url), 'lxml')
        page_class = parse_page(page_soup)

        print('found {} valid classes'.format(len(page_class)))
        append_to_sheet(page_class, grade_name)
        count += len(page_class)
        # show(page_class)
        # page_class_list.append(page_class)

        # sleep for 5 seconds
        time.sleep(5)

    # finish
    print('\nfinished crawling {} with {} classes in total\n\n'.format(grade_name, count))


def append_to_sheet(rows, sheet_name):
    global my_excel
    if sheet_name not in my_excel.sheetnames:
        sheet = my_excel.create_sheet(sheet_name)
        sheet.append(header_row)
    else:
        sheet = my_excel.get_sheet_by_name(sheet_name)

    for r in rows:
        sheet.append(r)
    my_excel.save(excel_name)
    print('append {} classes into sheet {}'.format(len(rows), sheet_name))


# define a method to show results
def show(a):
    pprint(a)


# using webdriver to load page
def myget(url):
    nobrowser = webdriver.PhantomJS()
    nobrowser.get(url)
    return nobrowser.page_source

#---------- main process ----------------
# for each grade, crawl its all classes
# store classes for each grade as a list
# store all those lists into a list
# speiyou = []
for g_url in grade_urls:
    # print('\n\n Crawl grade {}'.format(g_url))
    crawl_grade(g_url)  # use crawl_grade method as a blackbox
    # speiyou.append(grade_class_list)


# with open('all.txt', 'w') as f:
#     f.write(str(speiyou))






